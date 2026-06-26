"""
Export the breakeven total-return data to Excel / CSV for hand-replication.

Two products:
  export_full(path)            -- EVERYTHING used, day-level (business days), one sheet per
                                  tenor: which TIPS/UST CUSIP, their clean price & yield,
                                  accrued, index ratio, dirty, cash value V, DV01, the monthly
                                  DV01 denominator, daily dV/coupon/financing/PnL, repo rate,
                                  per-leg bp return, breakeven, and cumulatives. Plus a macro
                                  sheet (CPI, repo) and a README sheet with the formulas.
  export_returns(path, xT, xU) -- compact: just the returns (+ long/short BE at the given
                                  repo half-spreads), one sheet per tenor. Fast (from cache).

Replication chain (every column is in the sheet):
  dirty_real = clean + accrued ;  V = dirty_real * IR  (UST: IR = 1)
  financing  = days/360 * gc_repo/100 * V_prev
  PnL        = dV + coupon - financing            (dV = V - V_prev, same bond)
  bp         = PnL / denom                        (denom = month's 100k-DV01 rebalance)
  r_BE_bp    = r_TIPS_bp - r_UST_bp ;  cum_* = running sum (linear, not compounded)

Both entry points first call engine.refresh() — pull the latest data from Bloomberg
(data_layer.update) and rebuild returns_<tenor>.parquet — so the export is never stale.
Pass --no-update to skip the Bloomberg pull and export straight from the current cache
(also the automatic fallback if the Terminal isn't running).

Usage:
  python export.py                 # refresh, then -> exports/breakeven_full.xlsx + per-tenor CSVs
  python export.py returns 3 3     # refresh, then -> exports/breakeven_returns.xlsx at x_TIPS=x_UST=3bp
  python export.py --no-update     # export from the cached data as-is (no Bloomberg pull)
"""
from __future__ import annotations
import os, sys
import pandas as pd

import engine
import energy
import hedge

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE = engine.CACHE
EXPORTS = os.path.join(HERE, "exports")
TENORS = ["5y", "10y", "30y"]
# Hedge window for the per-tenor PnL columns. Desk decision: use the STATIC full-sample ratio
# (hedge.FULL_HEDGE) — one regression over the whole horizon, no monthly rebalance. Set to "2y" or
# "5y" for the rolling walk-forward ratio instead. (The hedge SHEET still lists the 2y/5y monthly
# ratios for reference.) The dashboard exposes a live 2Y/5Y/Full toggle.
EXPORT_HEDGE_WINDOW = "full"

# Output column order (per-tenor sheet), built in tenor_full().
FULL_COLS = [
    "settlement_date", "d", "gc_repo",
    # --- TIPS leg ---
    "TIPS_cusip", "TIPS_notional", "TIPS_DV01", "V_tips", "V_tips_prev", "TIPS_dirty_real", "TIPS_IR",
    "TIPS_clean", "TIPS_gross_bp", "TIPS_fin_bp", "r_TIPS_bp",
    # --- UST (nominal) leg ---
    "UST_cusip", "UST_notional", "UST_DV01", "V_nom", "V_nom_prev", "UST_dirty", "UST_clean",
    "UST_gross_bp", "UST_fin_bp", "r_UST_bp",
    # --- breakeven ---
    "net_financing_bp", "r_BE_bp", "cum_TIPS_bp", "cum_UST_bp", "cum_BE_bp",
    # --- flags ---
    "is_roll_day", "is_coupon_day", "is_weekend_or_holiday_step",
    "Is_5y_auction_date", "Is_10y_auction_date", "Is_30y_auction_date", "auction_size_bn",
]

README_ROWS = [
    ("date", "observation day. Marked to its T+1 settlement (settlement_date)."),
    ("settlement_date", "T+1 settle (bond-market calendar). V/accrued/IR are valued to THIS date."),
    ("d", "settlement-span = settle(t) - settle(t-1) in calendar days. Lands the weekend/holiday "
          "carry on the day BEFORE the weekend: FRIDAY d=3 (4 before a holiday), Mon d=1; a "
          "holiday's stale obs gets d=0. Accrual (via dV, settlement-date IR/accrued) AND repo "
          "both scale with the SAME d on the SAME day."),
    ("V_tips_prev / V_nom_prev", "prior settle value (the financing base): fin = d/360*gc/100*V_prev"),
    ("gc_repo", "GC financing rate that day (%, GCFRTSY; USRG1T/fed funds before 2009)"),
    ("TIPS_cusip / UST_cusip", "on-the-run bond each leg tracks (issue-date-gated TIPS-clock roll)"),
    ("*_notional", "face giving 100k DV01 = 1e7 / DV01 (set at the monthly reset, held all month)"),
    ("*_DV01", "sizing DV01 per 100 face (our calc, NOT BBG TIPS risk which is ~half); set at the "
               "monthly reset and held constant -> it is the bp denominator"),
    ("V_tips / V_nom", "cash value per 100 face: TIPS = dirty_real * IR ; UST = dirty"),
    ("*_dirty_real / UST_dirty", "clean + accrued (real for TIPS)"),
    ("TIPS_IR", "index ratio = DRI/base CPI (from CPI, matches Treasury to 1e-6); UST IR = 1"),
    ("*_clean", "quoted clean price per 100 (Bloomberg PX_CLEAN_MID)"),
    ("*_gross_bp", "leg price+coupon return before financing = (dV + coupon) / DV01, "
                   "dV = V(settle t) - V(settle t-1) so weekend accretion lands on Friday"),
    ("*_fin_bp", "leg financing drag in bp at GC mid = (d/360 * gc/100 * V_prev) / DV01"),
    ("r_TIPS_bp / r_UST_bp", "leg net daily return in bp = gross_bp - fin_bp"),
    ("net_financing_bp", "TIPS_fin_bp - UST_fin_bp (long TIPS pays, short UST earns; GC mid, x=0)"),
    ("r_BE_bp", "net breakeven daily return = r_TIPS_bp - r_UST_bp"),
    ("cum_*", "running LINEAR sum of daily bp (not compounded)"),
    ("is_roll_day", "either leg switched CUSIP that day"),
    ("is_coupon_day", "a coupon paid on either leg that day"),
    ("is_weekend_or_holiday_step", "d > 1 (the step spans a weekend/holiday)"),
    ("Is_5y/10y/30y_auction_date", "True if a real 5y/10y/30y TIPS was auctioned that calendar "
                                   "day — new issue OR reopening; sub-$1bn contingency auctions and "
                                   "nominal auctions are NOT flagged. <=1 flagged day per tenor/month."),
    ("auction_size_bn", "Offering amount of THAT day's TIPS auction, in $bn (per ticker — the "
                        "single auctioned security, not an aggregate). Blank on non-auction days. "
                        "Real auctions only: sub-$1bn contingency/test auctions (e.g. the 2020-07-10 "
                        "$25mn 5y) are excluded, so they neither flag nor show a size."),
    ("NOTE 1", "repo bid/offer x & specialness NOT in these numbers (GC mid). For long-BE apply "
               "(GC+x_tips) on TIPS / (GC-x_nom) on UST via the dashboard; short-BE flips."),
    ("NOTE 2", "BBG tie-out: V is the dirty price at settlement_date, so set BBG settlement to "
               "settlement_date (e.g. Friday's row -> Monday settle). dV(Friday) then carries the "
               "3-day weekend accrual + IR; repo on Friday uses d=3. Weekly total is unchanged vs "
               "booking on Monday -- the fix re-dates the carry, it does not re-size it."),
]

# Energy-hedge sheet column docs (RBOB gasoline; see energy.py). Lives on its own sheet.
ENERGY_README_ROWS = [
    ("--- ENERGY sheet (RBOB gasoline hedge) ---", "Front-month RBOB gasoline futures. ONE "
        "front price series (XB1) + ONE bp return series, built on the intersection of gas and "
        "bond trading days. Source: Bloomberg XB1/XB2 Comdty PX_LAST. See energy.py."),
    ("XB1", "generic 1st RBOB future close that day (cents/gallon) = the held FRONT-month price"),
    ("XB2", "generic 2nd RBOB future close that day (the contract that becomes the front next month)"),
    ("prior_mark", "the contract-consistent prior close used in the return: XB2(prev common day) "
                   "on a roll day (the new front's own close yesterday), else XB1(prev common day)"),
    ("days", "calendar-day gap to the prior COMMON trading day (2-3 across weekends/holidays). Gas "
             "days that aren't bond days are skipped; bond days that aren't gas days widen the gap "
             "(the return then spans 2-3 days), so the series lines up with the bond returns."),
    ("is_roll", "True on the 1st common trading day of a month -- the generic front rolled "
                "(XB1 expired end of prior month; yesterday's XB2 is today's XB1)."),
    ("chg", "XB1(t) - prior_mark (price points / cents-per-gallon). On a roll this is a clean "
            "single-contract move (XB1(t) - XB2(t-1)), never an old->new cross-contract jump."),
    ("r_bp", "daily gas return in bp = chg / prior_mark * 1e4 (percentage price return). THE "
             "return series; combine with the bond breakeven as r_BE_bp - h * r_bp at a hedge ratio h."),
    ("cum_bp", "running LINEAR sum of r_bp (not compounded), matching the bond cum_* convention."),
    ("ENERGY NOTE", "No financing/carry term: a same-contract futures return already excludes the "
                    "roll yield (curve carry = the XB1-XB2 spread, which the splice steps over). "
                    "Use `chg` to re-normalize to a $/contract (x420 for RBOB) or DV01 basis."),
    ("--- on each TENOR sheet ---", "Every energy column above also appears on the 5y/10y/30y "
        "sheets suffixed '_gas' (XB1_gas, usd_per_contract_gas, ...), aligned to the bond days "
        "(a bond day with no gas obs -> NaN; that move lands on the next common day). Plus the "
        "gasoline-hedge columns below (window = the `hedge` sheet's both-window detail, condensed "
        "here to a single selected window shown in `gas_hedge_window`):"),
    ("gas_hedge_window", "which lookback the per-tenor hedge ratios/PnL below use: 'full' = the "
        "STATIC full-sample ratio (one regression over the whole horizon, hardcoded in hedge.FULL_HEDGE, "
        "no rebalance — desk default); '2y'/'5y' = the trailing rolling walk-forward ratio."),
    ("gas_hedge_contracts_BE100 / _BE75", "the gasoline hedge ratio for THIS tenor, in # of 42,000-gal "
        "contracts (constant when window='full'; monthly when '2y'/'5y'). From a daily regression of "
        "breakeven $ on gas $/contract. BE100 = pure breakeven (β=1); BE75 = 75%-beta breakeven "
        "(TIPS − 0.75·UST). Positive => co-moving; SHORT this many vs a LONG BE. Full-sample β=100 "
        "ratios ≈ 5y 75.9, 10y 48.7, 30y 33.7 contracts."),
    ("r_BE75_bp", "75%-beta breakeven daily return = r_TIPS_bp − 0.75·r_UST_bp (vs r_BE_bp = β=100%)."),
    ("r_BE100_hedged_bp / r_BE75_hedged_bp", "gas-HEDGED daily breakeven return (bp) = BE − "
        "h·(gas $/contract)/$100k, where h is that month's BE100 / BE75 contract ratio. Months "
        "before the first ratio (window not full yet) are unhedged (h=0)."),
    ("cum_BE100_hedged_bp / cum_BE75_bp / cum_BE75_hedged_bp", "running LINEAR sums (bp) of the "
        "hedged BE100, the unhedged BE75, and the hedged BE75 daily returns (cf. cum_BE_bp = "
        "unhedged BE100). The hedge SHEET carries the full monthly ratios + R² for both betas/windows."),
]


_AUCT_INFO = None
def _auction_calendar_info():
    """TIPS-auction calendar for the export flags. Returns (date_sets, sizes):
      date_sets[tenor] = set of normalized REAL TIPS auction dates (new issues AND reopenings;
                         contingency/test auctions are excluded via auctions.real_tips_auctions,
                         e.g. the 2020-07-10 $25mn 5y reopening).
      sizes[date]      = that day's TIPS auction offering amount ($, per ticker -- NOT aggregated).
    Within the 2011+ analysis window every month carries exactly one real TIPS auction, so the
    per-tenor Is_* flag is unambiguous (<=1 flagged day/tenor/month) and no 'multi-auction month'
    review marker is needed -- the 2006-2010 doubles are pre-window and 2020-07's stray is dropped."""
    global _AUCT_INFO
    if _AUCT_INFO is None:
        import auctions
        a = auctions.real_tips_auctions().dropna(subset=["auctionDate"]).copy()
        a["nd"] = pd.to_datetime(a["auctionDate"]).dt.normalize()
        date_sets = {ten: set(a[a["tenor"] == ten]["nd"]) for ten in ("5y", "10y", "30y")}
        sizes = {pd.Timestamp(d): float(v) for d, v in zip(a["nd"], a["offeringAmount"]) if pd.notna(v)}
        _AUCT_INFO = (date_sets, sizes)
    return _AUCT_INFO


def tenor_full(tenor):
    """Assemble the full day-level reproducibility table for one tenor (both legs side by side,
    plus net breakeven + flags). Every column needed to hand-check a day on Bloomberg."""
    cpi = engine._macro()["cpi_nsa"]
    gc = engine.gc_series()
    t = engine.leg_series("tips", tenor, cpi, gc).add_prefix("TIPS_")
    u = engine.leg_series("nominal", tenor, cpi, gc).add_prefix("UST_")
    df = pd.concat([t, u], axis=1, sort=True)   # keep the row index date-sorted (pin vs pandas-4 default flip)
    out = pd.DataFrame(index=df.index)
    out.index.name = "date"
    out["settlement_date"] = df["TIPS_settle"].combine_first(df.get("UST_settle"))
    out["d"] = df["TIPS_days"].combine_first(df.get("UST_days"))
    out["gc_repo"] = df["TIPS_gc"].combine_first(df.get("UST_gc"))
    # TIPS leg
    out["TIPS_cusip"] = df["TIPS_cusip"]; out["TIPS_notional"] = df["TIPS_notional"]
    out["TIPS_DV01"] = df["TIPS_denom"]; out["V_tips"] = df["TIPS_V"]; out["V_tips_prev"] = df["TIPS_V_prev"]
    out["TIPS_dirty_real"] = df["TIPS_dirty_real"]; out["TIPS_IR"] = df["TIPS_IR"]
    out["TIPS_clean"] = df["TIPS_clean"]; out["TIPS_gross_bp"] = df["TIPS_gross_bp"]
    out["TIPS_fin_bp"] = df["TIPS_fin_bp"]; out["r_TIPS_bp"] = df["TIPS_bp"]
    # UST leg
    out["UST_cusip"] = df["UST_cusip"]; out["UST_notional"] = df["UST_notional"]
    out["UST_DV01"] = df["UST_denom"]; out["V_nom"] = df["UST_V"]; out["V_nom_prev"] = df["UST_V_prev"]
    out["UST_dirty"] = df["UST_dirty_real"]; out["UST_clean"] = df["UST_clean"]
    out["UST_gross_bp"] = df["UST_gross_bp"]; out["UST_fin_bp"] = df["UST_fin_bp"]
    out["r_UST_bp"] = df["UST_bp"]
    # breakeven
    out["net_financing_bp"] = df["TIPS_fin_bp"] - df["UST_fin_bp"]
    out["r_BE_bp"] = df["TIPS_bp"] - df["UST_bp"]
    out["cum_TIPS_bp"] = out["r_TIPS_bp"].cumsum()
    out["cum_UST_bp"] = out["r_UST_bp"].cumsum()
    out["cum_BE_bp"] = out["r_BE_bp"].cumsum()
    # flags
    f = lambda s: s.fillna(False).astype(bool)
    out["is_roll_day"] = f(df.get("TIPS_is_roll")) | f(df.get("UST_is_roll"))
    out["is_coupon_day"] = f(df.get("TIPS_is_coupon")) | f(df.get("UST_is_coupon"))
    out["is_weekend_or_holiday_step"] = out["d"] > 1
    # auction flags (a real TIPS of that tenor auctioned that day: new issue OR reopening; contingency
    # auctions excluded) + that day's offering size ($bn). Same calendar columns in every tenor sheet.
    asets, sizes = _auction_calendar_info()
    nd = out.index.normalize()
    for ten in ("5y", "10y", "30y"):
        out[f"Is_{ten}_auction_date"] = nd.isin(asets[ten])
    out["auction_size_bn"] = [round(sizes[d] / 1e9, 3) if d in sizes else None for d in nd]
    out = out.reindex(columns=FULL_COLS)
    # --- fold in the gasoline hedge: energy columns (suffixed _gas) + the month's contract
    # hedge ratio for THIS tenor (held all month, in sync with the DV01 rebalance). Energy is
    # aligned to the bond days; a bond day with no gas obs (gas-holiday span, e.g. Good Friday)
    # shows NaN gas -- that day's gas move lands on the next common day, by construction.
    nrg = energy_full()
    if nrg is not None:
        g = nrg.reindex(out.index)
        for col in nrg.columns:
            out[f"{col}_gas"] = g[col]
        # gasoline hedge (selected window EXPORT_HEDGE_WINDOW): contract ratio for the pure
        # breakeven (BE100, β=1) and the 75%-beta breakeven (BE75, β=0.75), held each month, plus
        # the gas-HEDGED daily/cum P&L. Hedged BE = BE − h·(gas $/contract)/$100k  [short h contracts
        # vs long BE]. Early months with no ratio yet are unhedged (h treated as 0 for the P&L).
        W = EXPORT_HEDGE_WINDOW
        gas = hedge.daily_gas_usd(out.index)                       # 0-filled daily gas $/contract
        h100 = hedge.daily_hedge_beta(tenor, out.index, W, 1.0)
        h75 = hedge.daily_hedge_beta(tenor, out.index, W, 0.75)
        out["gas_hedge_window"] = W
        out["gas_hedge_contracts_BE100"] = h100
        out["gas_hedge_contracts_BE75"] = h75
        out["r_BE75_bp"] = out["r_TIPS_bp"] - 0.75 * out["r_UST_bp"]
        out["r_BE100_hedged_bp"] = out["r_BE_bp"] - h100.fillna(0.0) * gas / 1e5
        out["r_BE75_hedged_bp"] = out["r_BE75_bp"] - h75.fillna(0.0) * gas / 1e5
        out["cum_BE100_hedged_bp"] = out["r_BE100_hedged_bp"].cumsum()
        out["cum_BE75_bp"] = out["r_BE75_bp"].cumsum()
        out["cum_BE75_hedged_bp"] = out["r_BE75_hedged_bp"].cumsum()
    return out


ENERGY_COLS = ["XB1", "XB2", "prior_mark", "days", "is_roll", "chg", "usd_per_contract", "r_bp", "cum_bp"]


HEDGE_COLS = ["tenor",
              "BE100_2y", "BE100_2y_R2", "BE100_5y", "BE100_5y_R2",
              "BE75_2y", "BE75_2y_R2", "BE75_5y", "BE75_5y_R2", "n_2y", "n_5y"]


def hedge_full():
    """Monthly walk-forward hedge-ratio table per tenor — BOTH the pure breakeven (BE100, β=1.0)
    and the 75%-beta breakeven (BE75, β=0.75), each for the 2y and 5y windows, with R². One row
    per (tenor, rebalance month). None if unavailable. See hedge.py."""
    frames = []
    for ten in TENORS:
        try:
            a = hedge.monthly_hedge_ratios(ten, beta=1.0)     # BE100
            b = hedge.monthly_hedge_ratios(ten, beta=0.75)    # BE75 (same rebalance months/order)
        except Exception as e:
            print(f"  hedge sheet {ten} skipped ({type(e).__name__}: {e})")
            continue
        if a.empty:
            continue
        m = pd.DataFrame({"tenor": ten}, index=a["month"].dt.to_period("M").astype(str))
        for w in ("2y", "5y"):
            m[f"BE100_{w}"] = a[f"hedge_contracts_{w}"].to_numpy()
            m[f"BE100_{w}_R2"] = a[f"r2_{w}"].to_numpy()
            m[f"BE75_{w}"] = b[f"hedge_contracts_{w}"].to_numpy()
            m[f"BE75_{w}_R2"] = b[f"r2_{w}"].to_numpy()
            m[f"n_{w}"] = a[f"n_{w}"].to_numpy()
        frames.append(m)
    if not frames:
        print("  hedge sheet skipped — no data")
        return None
    df = pd.concat(frames).reindex(columns=HEDGE_COLS)
    df.index.name = "rebalance_month"
    return df


def energy_full():
    """The energy-hedge day-level frame (RBOB gasoline) for the export, or None if unavailable.
    Loads the built series; falls back to building from the raw cache if the parquet is missing."""
    try:
        df = energy.load_energy()
    except Exception:
        try:
            df = energy.build_energy(save=False)
        except Exception as e:
            print(f"  energy sheet skipped — no data ({type(e).__name__}: {e})")
            return None
    df = df.reindex(columns=ENERGY_COLS)
    df.index.name = "date"
    return df


def _format_sheet(ws, date_col=True):
    """Fix the cosmetics: date column shows as YYYY-MM-DD (not a too-wide datetime that
    renders as #######), sensible column widths, bold + frozen header row."""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    ws.freeze_panes = "B2"
    for cell in ws[1]:                                  # header row
        cell.font = Font(bold=True)
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        header = str(col[0].value or "")
        ws.column_dimensions[letter].width = max(12, len(header) + 2)
    if date_col:                                        # column A = the date index
        ws.column_dimensions["A"].width = 12
        for cell in ws["A"][1:]:
            cell.number_format = "yyyy-mm-dd"


def export_full(path=None):
    os.makedirs(EXPORTS, exist_ok=True)
    path = path or os.path.join(EXPORTS, "breakeven_full.xlsx")
    frames = {}
    for ten in TENORS:
        print(f"  building {ten} ...", flush=True)
        frames[ten] = tenor_full(ten)
    macro = engine._macro()
    nrg = energy_full()                                    # energy-hedge sheet (RBOB gasoline); None if unavailable
    hdg = hedge_full()                                      # monthly contract hedge-ratio table; None if unavailable
    readme = pd.DataFrame(README_ROWS + (ENERGY_README_ROWS if nrg is not None else []),
                          columns=["column", "description"])
    locked = []
    try:                                                   # the multi-sheet workbook
        with pd.ExcelWriter(path, engine="openpyxl") as xl:
            readme.to_excel(xl, sheet_name="README", index=False)
            _format_sheet(xl.sheets["README"], date_col=False)
            for ten in TENORS:
                frames[ten].to_excel(xl, sheet_name=ten)
                _format_sheet(xl.sheets[ten])
            if nrg is not None:
                nrg.to_excel(xl, sheet_name="energy")
                _format_sheet(xl.sheets["energy"])
            if hdg is not None:
                hdg.to_excel(xl, sheet_name="hedge")
                _format_sheet(xl.sheets["hedge"], date_col=False)
            macro.to_excel(xl, sheet_name="macro")
            _format_sheet(xl.sheets["macro"])
        print(f"  wrote {path}  ({', '.join(TENORS)} sheets"
              f"{' + energy' if nrg is not None else ''}{' + hedge' if hdg is not None else ''}"
              f" + macro + README)")
    except PermissionError:
        locked.append(os.path.basename(path))
    for ten in TENORS:                                     # per-tenor CSVs (each independent)
        try:
            frames[ten].to_csv(os.path.join(EXPORTS, f"breakeven_{ten}.csv"))
        except PermissionError:
            locked.append(f"breakeven_{ten}.csv")
    if nrg is not None:                                    # standalone energy CSV
        try:
            nrg.to_csv(os.path.join(EXPORTS, "energy.csv"))
        except PermissionError:
            locked.append("energy.csv")
    if hdg is not None:                                    # standalone hedge-ratio CSV
        try:
            hdg.to_csv(os.path.join(EXPORTS, "hedge_ratios.csv"))
        except PermissionError:
            locked.append("hedge_ratios.csv")
    if locked:
        print(f"  !! could not write (file open/locked — close it and re-run): {', '.join(locked)}")
    else:
        print(f"  wrote per-tenor CSVs in {EXPORTS}")
    return path


def export_returns(path=None, xT=0.0, xU=0.0):
    """Compact returns export at given repo half-spreads (used by the interactive button)."""
    os.makedirs(EXPORTS, exist_ok=True)
    path = path or os.path.join(EXPORTS, "breakeven_returns.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        for ten in TENORS:
            p = os.path.join(CACHE, f"returns_{ten}.parquet")
            if not os.path.exists(p):
                continue
            d = engine.apply_spread(engine.load_returns(ten), xT, xU)
            d["cum_longBE_bp"] = d["longBE_bp"].cumsum()
            d["cum_shortBE_bp"] = d["shortBE_bp"].cumsum()
            d["cum_BEmid_bp"] = d["BEmid_bp"].cumsum()
            d.index.name = "date"
            d.to_excel(xl, sheet_name=ten)
            _format_sheet(xl.sheets[ten])
    print(f"  wrote {path}  (x_TIPS={xT}bp, x_UST={xU}bp; sheets {', '.join(TENORS)})")
    return path


if __name__ == "__main__":
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding="utf-8")
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    update = "--no-update" not in sys.argv
    engine.refresh(update_data=update)                          # fresh bond data + returns first
    energy.refresh(update_data=update)                          # + energy hedge (XB1/XB2 pull + build)
    try:
        hedge.build_all()                                       # + walk-forward contract hedge ratios
    except Exception as e:
        print(f"  [export] hedge-ratio build skipped: {type(e).__name__}: {e}")
    if args and args[0] == "returns":
        xT = float(args[1]) if len(args) > 1 else 0.0
        xU = float(args[2]) if len(args) > 2 else xT
        export_returns(xT=xT, xU=xU)
    else:
        export_full()

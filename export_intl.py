"""
Export European/UK linker financed total returns to Excel/CSV — ONE SHEET PER BOND.

Mirrors export.py (the US hand-replication dump) but for the non-US linkers, and laid out one
worksheet per ISIN (the desk's ask), plus a UNIVERSE sheet (catalog + per-market conventions), a
macro sheet (reference indices + financing) and a README sheet (column formulas). Outright
financed total return only — no breakeven leg / gas hedge yet (reference_intl.MD scope).

Replication chain (every column is on the sheet):
  IR        = DRI(settle)/DRI(dated)        [rules-based; IR_bbg = Bloomberg's, for cross-check]
  V         = (clean + accrued) * IR
  financing = days/360 * gc/100 * V_prev    (local GC: €STR euro / SONIA gilt)
  pnl       = dV + coupon - financing       (dV = V - V_prev, same bond; coupon = C/freq * IR)
  bp        = pnl / DV01_denom              (denom = month's 100k-DV01 rebalance)
  cum_bp    = running LINEAR sum

Usage:
  python export_intl.py                 # refresh (pull + rebuild) -> exports/linkers_returns.xlsx + per-bond CSVs
  python export_intl.py --no-update     # export straight from cache_intl (no Bloomberg pull)
"""
from __future__ import annotations
import os, sys
import pandas as pd

import linkers
import engine_intl as eng
import data_layer_intl as dl

HERE = os.path.dirname(os.path.abspath(__file__))
EXPORTS = os.path.join(HERE, "exports")

SHEET_COLS = ["settlement_date", "d", "gc_repo", "market", "country", "program",
              "clean", "yield", "accrued", "IR", "IR_bbg", "dirty_real", "V", "V_prev",
              "DV01", "notional", "dV", "coupon", "gross_bp", "fin_bp", "bp", "cum_bp", "is_coupon"]

README_ROWS = [
    ("date", "observation day. Marked to its T+settle business day (settlement_date)."),
    ("settlement_date", "T+1 (gilt) or T+2 (euro) settle on the local calendar; V/accrued/IR valued here."),
    ("d", "settlement span = settle(t)-settle(t-1) in calendar days; drives accrual (via dV) and repo together."),
    ("gc_repo", "local GC financing rate that day (%, €STR for euro, SONIA for gilts)"),
    ("market/country/program", "linkers.MARKETS code, e.g. IT_BTPEI / IT / BTP€i"),
    ("clean", "quoted clean REAL price per 100 (Bloomberg PX_CLEAN_MID)"),
    ("yield", "real yield to maturity (Bloomberg YLD_YTM_MID)"),
    ("accrued", "accrued REAL coupon per 100 = (C/freq)*(1-w), act/act ICMA"),
    ("IR", "index ratio = DRI(settle)/DRI(dated), rules-based from the reference index (reference_intl §3)"),
    ("IR_bbg", "Bloomberg's own INDEX_RATIO that day — cross-check vs IR; engine drives off IR"),
    ("dirty_real", "clean + accrued (real)"),
    ("V", "cash value per 100 face = dirty_real * IR (the inflation-uplifted settlement amount)"),
    ("V_prev", "prior-settle V (the financing base): financing = d/360 * gc/100 * V_prev"),
    ("DV01", "sizing DV01 per 100 face (our pricing.py calc); set at the monthly reset, the bp denominator"),
    ("notional", "face giving 100k DV01 = 1e7/DV01 (set at monthly reset, held all month)"),
    ("dV", "V(settle t) - V(settle t-1), same bond (weekend accretion lands on the pre-weekend day)"),
    ("coupon", "coupon cash booked when a pay date falls in the settle span = (C/freq)*IR"),
    ("gross_bp", "(dV + coupon)/DV01 — price+coupon return before financing"),
    ("fin_bp", "financing drag in bp at GC mid = (d/360*gc/100*V_prev)/DV01"),
    ("bp", "net daily financed return in bp = gross_bp - fin_bp"),
    ("cum_bp", "running LINEAR sum of bp (not compounded)"),
    ("is_coupon", "a coupon paid that day"),
    ("NOTE financing", "local funding per bond (no FX / no cross-currency normalization yet — desk decision pending)."),
    ("NOTE floor", "deflation par floor (euro markets) affects only the redemption mark and is in the "
                   "market price; low-coupon bonds with IR<1 are flagged in the UNIVERSE sheet."),
    ("NOTE breakeven", "outright real-bond return only; no nominal-leg/breakeven yet (nominal comparator TBD)."),
]


def bond_sheet(isin, market):
    """One bond's export frame (rebuilds from cache if the returns parquet is absent)."""
    try:
        df = eng.load_returns(isin)
    except Exception:
        df = eng.build_bond(isin, market, save=False)
    if df.empty:
        return None
    c = linkers.conv(market)
    out = df.rename(columns={"gc": "gc_repo", "settle": "settlement_date", "days": "d"}).copy()
    out["country"] = c["country"]
    out["program"] = c["program"]
    out.index.name = "date"
    return out.reindex(columns=SHEET_COLS)


def universe_sheet():
    """Catalog + per-market conventions for the README/UNIVERSE sheet."""
    u = linkers.load_universe(include_deferred=True)
    mm = linkers.market_matrix().reset_index()
    return u, mm


def _format(ws, date_col=True):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    ws.freeze_panes = "B2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = max(11, len(str(col[0].value or "")) + 2)
    if date_col:
        ws.column_dimensions["A"].width = 12
        for cell in ws["A"][1:]:
            cell.number_format = "yyyy-mm-dd"


def export(path=None, include_deferred=False):
    os.makedirs(EXPORTS, exist_ok=True)
    path = path or os.path.join(EXPORTS, "linkers_returns.xlsx")
    u = linkers.load_universe(include_deferred=include_deferred)
    sheets, empties = {}, []
    for _, r in u.iterrows():
        s = bond_sheet(r["isin"], r["market"])
        (sheets.__setitem__(r["isin"], s) if s is not None else empties.append(r["isin"]))
    uni, mm = universe_sheet()
    readme = pd.DataFrame(README_ROWS, columns=["column", "description"])
    try:
        macro = dl._macro()
    except Exception:
        macro = None
    locked = []
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as xl:
            readme.to_excel(xl, sheet_name="README", index=False); _format(xl.sheets["README"], date_col=False)
            uni.to_excel(xl, sheet_name="UNIVERSE", index=False); _format(xl.sheets["UNIVERSE"], date_col=False)
            mm.to_excel(xl, sheet_name="CONVENTIONS", index=False); _format(xl.sheets["CONVENTIONS"], date_col=False)
            if macro is not None:
                macro.to_excel(xl, sheet_name="macro"); _format(xl.sheets["macro"])
            for isin, s in sheets.items():
                s.to_excel(xl, sheet_name=isin[:31]); _format(xl.sheets[isin[:31]])
        print(f"  wrote {path}  ({len(sheets)} bond sheets + UNIVERSE/CONVENTIONS/macro/README)")
    except PermissionError:
        locked.append(os.path.basename(path))
    csvdir = os.path.join(EXPORTS, "linkers")
    os.makedirs(csvdir, exist_ok=True)
    for isin, s in sheets.items():
        try:
            s.to_csv(os.path.join(csvdir, f"{isin}.csv"))
        except PermissionError:
            locked.append(f"{isin}.csv")
    if empties:
        print(f"  no data (skipped sheets): {len(empties)} — {', '.join(empties[:8])}{'...' if len(empties) > 8 else ''}")
    if locked:
        print(f"  !! could not write (file open/locked): {', '.join(locked)}")
    return path


if __name__ == "__main__":
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding="utf-8")
    update = "--no-update" not in sys.argv
    eng.refresh(update_data=update)
    export()
